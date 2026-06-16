"""Load TV live transmission orders into dashboard database."""
import openpyxl, os, sqlite3
from datetime import datetime
from load_payout import fmt_date, fetch_biz_types, DB_PATH, BQ_KEY

EXCEL_PATH = os.path.expanduser("~/Downloads/payout.xlsx")
PERIOD_LABEL = "25 - 31 May 2026"
PAYMENT_DATE = "10/06/2026"
SORT_DATE = "2026-05-31"
CHANNEL = "tv"


def load_tv():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print(f"Sheet: {ws.title}, Rows: {ws.max_row}")

    # Parse: A=Order id, B=Product name, C=Status, E=Amount, F=Seller
    orders_by_store = {}
    all_short_ids = []
    total_orders = 0

    for i in range(2, ws.max_row + 1):
        order_id = ws.cell(row=i, column=1).value
        product = ws.cell(row=i, column=2).value
        status = ws.cell(row=i, column=3).value
        amount = ws.cell(row=i, column=5).value
        seller = ws.cell(row=i, column=6).value
        if not order_id or not seller:
            continue

        seller = str(seller).strip()
        order_id = str(order_id).strip()
        product = str(product).strip() if product else "-"
        status = str(status).strip() if status else "-"

        is_paid = False
        if amount and isinstance(amount, (int, float)):
            price = float(amount)
        elif amount and str(amount).strip().lower() == "paid":
            price = 0
            is_paid = True
        else:
            price = 0

        total_orders += 1
        all_short_ids.append(order_id)
        if seller not in orders_by_store:
            orders_by_store[seller] = []
        orders_by_store[seller].append({
            "bpid": order_id,
            "product_name": product,
            "status": status,
            "price": price,
            "is_paid": is_paid,
            "category": "-",
            "customer": "-",
            "user_city": "-",
            "quantity": 1,
            "commission": 0,
            "delivery_date": "-",
            "order_date": "-",
        })

    print(f"Parsed: {total_orders} orders across {len(orders_by_store)} sellers")

    # Fetch from BigQuery using short IDs (last 4 chars of buypassid)
    print(f"Fetching details for {len(all_short_ids)} orders from BigQuery...")
    try:
        from google.cloud import bigquery
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = BQ_KEY
        client = bigquery.Client(project="buypass-ai")

        bq_map = {}
        for i in range(0, len(all_short_ids), 200):
            batch = all_short_ids[i:i + 200]
            ids_str = ", ".join(f"'{sid}'" for sid in batch)
            query = f"""
                SELECT buypassid, userdetails_name, userdetails_city, createdon, modifiedon, status
                FROM `buypass-ai.BuyPASS.orders`
                WHERE SUBSTR(buypassid, -4) IN ({ids_str})
            """
            for row in client.query(query).result():
                short = row.buypassid[-4:]
                bq_map[short] = {
                    "full_bpid": row.buypassid,
                    "customer": row.userdetails_name or "-",
                    "user_city": row.userdetails_city or "-",
                    "order_date": row.createdon,
                    "delivery_date": row.modifiedon,
                }
            print(f"  Batch {i+1}-{i+len(batch)}: {len(bq_map)} matched")

        print(f"BQ orders matched: {len(bq_map)} / {len(all_short_ids)}")

        # Fetch product names
        product_map = {}
        full_bpids = [v["full_bpid"] for v in bq_map.values()]
        for i in range(0, len(full_bpids), 200):
            batch = full_bpids[i:i + 200]
            ids_str = ", ".join(f"'{bid}'" for bid in batch)
            query = f"""
                SELECT buypassid, nameen
                FROM `buypass-ai.BuyPASS.order_details`
                WHERE buypassid IN ({ids_str})
            """
            for row in client.query(query).result():
                product_map[row.buypassid] = row.nameen or "-"
            print(f"  Products batch {i+1}-{i+len(batch)}")

        # Enrich orders
        for store, orders in orders_by_store.items():
            for o in orders:
                bq = bq_map.get(o["bpid"])
                if bq:
                    o["customer"] = bq["customer"]
                    o["user_city"] = bq["user_city"]
                    o["order_date"] = fmt_date(bq["order_date"]) if bq["order_date"] else "-"
                    o["delivery_date"] = fmt_date(bq["delivery_date"]) if bq["delivery_date"] else "-"
                    bq_product = product_map.get(bq["full_bpid"])
                    if bq_product and bq_product != "-":
                        o["product_name"] = bq_product

    except Exception as e:
        print(f"WARNING: BigQuery fetch failed: {e}")

    # Build sellers
    sellers = []
    for store, orders in sorted(orders_by_store.items()):
        amount = sum(o["price"] for o in orders)
        sellers.append({
            "store": store,
            "amount": amount,
            "total_orders": len(orders),
            "biz_type": "-",
            "iban": "-",
            "bank": "-",
            "account_title": "-",
        })

    total_payout = sum(s["amount"] for s in sellers)
    print(f"\nSellers: {len(sellers)}, Total payout: Rs. {total_payout:,.0f}")

    # Fetch biz types
    biz_map = fetch_biz_types([s["store"] for s in sellers])
    for s in sellers:
        if s["store"] in biz_map:
            s["biz_type"] = biz_map[s["store"]]

    # Sample
    print("\nSample sellers:")
    for s in sellers[:5]:
        print(f"  {s['store']}: Rs.{s['amount']:,.0f} | {s['total_orders']} orders | {s['biz_type']}")

    sample_store = list(orders_by_store.keys())[0]
    print(f"\nSample orders ({sample_store}):")
    for o in orders_by_store[sample_store][:3]:
        paid = " [PAID]" if o["is_paid"] else ""
        print(f"  {o['bpid']}: {o['product_name'][:40]} | {o['customer']} | Ordered: {o['order_date']} | Rs.{o['price']}{paid}")

    # Save to SQLite
    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()

    cur.execute("SELECT id FROM periods WHERE period_label = ? AND channel = ?", (PERIOD_LABEL, CHANNEL))
    existing = cur.fetchone()
    if existing:
        pid = existing[0]
        cur.execute("DELETE FROM orders WHERE period_id = ?", (pid,))
        cur.execute("DELETE FROM sellers WHERE period_id = ?", (pid,))
        cur.execute("DELETE FROM periods WHERE id = ?", (pid,))
        print("\nReplaced existing TV period")

    upload_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
    cur.execute(
        "INSERT INTO periods (period_label, upload_time, total_sellers, total_orders, total_payout, payment_date, sort_date, channel) VALUES (?,?,?,?,?,?,?,?)",
        (PERIOD_LABEL, upload_time, len(sellers), total_orders, total_payout, PAYMENT_DATE, SORT_DATE, CHANNEL),
    )
    pid = cur.lastrowid

    for s in sellers:
        cur.execute(
            "INSERT INTO sellers (period_id, store, amount, total_orders, biz_type, iban, bank, account_title, payment_date, channel) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (pid, s["store"], s["amount"], s["total_orders"], s["biz_type"], s["iban"], s["bank"], s["account_title"], None, CHANNEL),
        )

    for store, orders in orders_by_store.items():
        for o in orders:
            # Store is_paid flag in commission column (1=paid, 0=unpaid)
            paid_flag = 1 if o["is_paid"] else 0
            cur.execute(
                "INSERT INTO orders (period_id, store, bpid, category, product_name, customer, user_city, quantity, price, commission, delivery_date, order_date, channel) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (pid, store, o["bpid"], o["category"], o["product_name"], o["customer"], o["user_city"], o["quantity"], o["price"], paid_flag, o["delivery_date"], o["order_date"], CHANNEL),
            )

    db.commit()
    db.close()
    print(f"\nDONE! TV period saved as period_id={pid}")
    return pid


if __name__ == "__main__":
    load_tv()
